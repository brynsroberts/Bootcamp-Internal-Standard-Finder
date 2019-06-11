"""
Name: Bryan Roberts
Date: 6/10/2019
Description: Program searches excel files in current directory of script for
Fiehn Lab HILIC internal standards.  Writes results to new sheet called "results.xlsx" 
that writes 'Y' if internal standards is foundand 'N' if internal standard is not 
found.  For untargeted metabolomics bootcamp paper focusing on MS-Dial output.
Sources:
https://automatetheboringstuff.com/
http://prime.psc.riken.jp/Metabolomics_Software/MS-DIAL/
Excel File Format:
Standard exported output from MS-Dial
"""

import openpyxl
import os.path

#return list of excel documents in folder
def getExcelSheets():
    excelSheets = []
    for file in os.listdir():
        if file[-5:] == '.xlsx':
            if file[0] != '~':
                excelSheets.append(os.path.join(os.getcwd(), file))
    return excelSheets

#return current filename
def getFileName(excelSheets, index):
    split = excelSheets[index].split(os.path.sep)
    return split[-1]

#opens and returns workbook of current excel sheet
def openWorkBook (excelSheets, index):
    wb = openpyxl.load_workbook(excelSheets[index])
    return wb

#returns sheet of current workbook
def makeSheet(wb):
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    return sheet

#makes new sheet called results and returns sheet
def makeResultsWorkBook():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Standard Name'
    currentRow = 2

    #write internal standard names to result sheet
    for name in standards:
        sheet.cell(row=currentRow, column=1).value = name
        currentRow += 1

    sheet.cell(row=currentRow + 1, column=1).value = 'Count'
    wb.save('results.xlsx')
    return wb

#finds standards and writes results to return sheet
def findStandards(sheet, results, currentRow, currentColumn):
    count = 0       #count number of internal standard found in each file
    found = False   #bool if individual standard is found, changes to true

    #top results row with filename
    results.cell(row = 1, column = currentColumn).value = fileName

    #loop through all internal standards in standard dictionary
    for name in standards:

        found = False   #reset found for each standard

        #check to find mz and rt match
        for rowNum in range(5, sheet.max_row):
            retentionTime = float(sheet.cell(row=rowNum, column=2).value)
            libraryRetentionTime = standards[name]['rt']

            #retention time match
            if retentionTime < (libraryRetentionTime + 0.05) and retentionTime > (libraryRetentionTime - 0.05):
                massToCharge = float(sheet.cell(row=rowNum, column=3).value)
                libraryMassToCharge = standards[name]['mz']

                #mz match
                if massToCharge < (libraryMassToCharge + 0.005) and massToCharge > (libraryMassToCharge - 0.005):
                    found = True
                    count += 1

        #write result to 'results.xlsx'
        if found:
            results.cell(row=currentRow, column=currentColumn).value = 'Y'
        else:
            results.cell(row=currentRow, column=currentColumn).value = 'N'
        currentRow += 1

    #print count
    results.cell(row=currentRow + 1, column=currentColumn).value = count


#Fiehn Lab HILIC Internal Standards    
standards = {'CUDA': {'mz': 341.2799, 'rt': 1.16},
             'D3-Creatinine': {'mz': 117.0850, 'rt': 4.95},
             'D9-Choline': {'mz': 113.1635, 'rt': 5.18},
             'D9-TMAO': {'mz': 85.1322, 'rt': 5.58},
             'D3-1-Methylnicotinamide': {'mz': 140.0898, 'rt': 6.26},
             'Val-Try-Val': {'mz': 380.2180, 'rt': 6.96},
             'D9-Betaine': {'mz': 127.1427, 'rt': 7.25},
             'D3-AC(2:0)': {'mz': 207.1419, 'rt': 7.21},
             'D3-Histamine N-methyl': {'mz': 129.1214, 'rt': 7.35},
             'D3-L-Carnitine': {'mz': 165.1313, 'rt': 7.82},
             'D9-Butyrobetaine': {'mz': 155.1740, 'rt': 7.82},
             'D9-Crotonobetaine': {'mz': 153.1584, 'rt': 7.86},
             'D3-Creatine': {'mz': 135.0956, 'rt': 8.15},
             'D3-Alanine': {'mz': 93.0738, 'rt': 8.17},
             'D5-L-Glutamine': {'mz': 152.1078, 'rt': 8.67},
             'D3-DL-Glutamic Acid': {'mz': 151.0793, 'rt': 8.85},
             'D3-DL-Aspartic Acid': {'mz': 137.0636, 'rt': 9.34},
             '15N2-L-Arginine': {'mz': 177.1130, 'rt': 9.53}
             }

"""
Execute main program
"""

if __name__ == "__main__":

    #initialize excelSheets list and open results file
    excelSheets = getExcelSheets()
    resultsWorkBook = makeResultsWorkBook()
    results = makeSheet(resultsWorkBook)
    
    #starting results row and column
    currentColumn = 2
    currentRow = 2

    #for each excel file found perform loop
    for index in range(len(excelSheets)):
        fileName = getFileName(excelSheets, index)
        wb = openWorkBook(getExcelSheets(), index)
        sheet = makeSheet(wb)
        findStandards(sheet, results, currentRow, currentColumn)
        resultsWorkBook.save('results.xlsx')
        
        #update results row and column
        currentColumn += 1
        currentRow = 2





