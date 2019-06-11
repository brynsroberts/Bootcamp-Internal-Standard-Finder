
import openpyxl
import os.path
from glob import glob
from xlrd import open_workbook

def getExcelSheets():
    excelSheets = []
    for file in os.listdir():
        if file[-5:] == '.xlsx':
            if file[0] != '~':
                excelSheets.append(os.path.join(os.getcwd(), file))
    return excelSheets

def getFileName(excelSheets, index):
    split = excelSheets[index].split(os.path.sep)
    return split[-1]

def openWorkBook (excelSheets, index):
    wb = openpyxl.load_workbook(excelSheets[index])
    return wb

def makeSheet(wb):
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    return sheet

def makeResultsSheet(wb, fileName):
    results = wb.create_sheet(index=1, title='results')
    results['A1'] = 'Standard Name'
    results['B1'] = fileName

    currentRow = 2
    for name in standards:
        results.cell(row=currentRow, column=1).value = name
        currentRow += 1

    results.cell(row=currentRow + 1, column=1).value = 'Count'
    return results

def findStandards(sheet, results, currentRow, currentColumn):
    count = 0
    found = False

    for name in standards:

        found = False

        for rowNum in range(5, sheet.max_row):
            retentionTime = float(sheet.cell(row=rowNum, column=2).value)
            libraryRetentionTime = standards[name]['rt']

            if retentionTime < (libraryRetentionTime + 0.05) and retentionTime > (libraryRetentionTime - 0.05):
                massToCharge = float(sheet.cell(row=rowNum, column=3).value)
                libraryMassToCharge = standards[name]['mz']

                if massToCharge < (libraryMassToCharge + 0.005) and massToCharge > (libraryMassToCharge - 0.005):
                    found = True
                    count += 1

        if found:
            results.cell(row=currentRow, column=currentColumn).value = 'Y'
        else:
            results.cell(row=currentRow, column=currentColumn).value = 'N'
        currentRow += 1

    results.cell(row=currentRow + 1, column=currentColumn).value = count

def saveAndClose(wb, results):
    wb.save('results.xlsx')
    
standards = {'CUDA': {'mz': 341.2799, 'rt': 1.16},
             'D3-Creatinine': {'mz': 117.0850, 'rt': 4.95},
             'D9-Choline': {'mz': 113.1635, 'rt': 5.18},
             'D9-TMAO': {'mz': 85.1322, 'rt': 5.58},
             'D3-1-Methylnicotinamide': {'mz': 140.0898, 'rt': 6.26},
             'Val-Try-Val': {'mz': 380.2180, 'rt': 6.96},
             'D9-Betaine': {'mz': 127.1427, 'rt': 7.25},
             'D3-AC(2:0)': {'mz': 207.1419, 'rt': 7.21},
             'D3-Histamine N-methyl': {'mz': 129.1214, 'rt': 7.35},
             'D3-L-Carnitine': {'mz': 165.1313, 'rt': 7.82},
             'D9-Butyrobetaine': {'mz': 155.1740, 'rt': 7.82},
             'D9-Crotonobetaine': {'mz': 153.1584, 'rt': 7.86},
             'D3-Creatine': {'mz': 135.0956, 'rt': 8.15},
             'D3-Alanine': {'mz': 93.0738, 'rt': 8.17},
             'D5-L-Glutamine': {'mz': 152.1078, 'rt': 8.67},
             'D3-DL-Glutamic Acid': {'mz': 151.0793, 'rt': 8.85},
             'D3-DL-Aspartic Acid': {'mz': 137.0636, 'rt': 9.34},
             '15N2-L-Arginine': {'mz': 177.1130, 'rt': 9.53}
             }

excelSheets = getExcelSheets()

for index in range(len(excelSheets)):
    fileName = getFileName(excelSheets, index)
    wb = openWorkBook(getExcelSheets(), index)
    sheet = makeSheet(wb)
    results = makeResultsSheet(wb, fileName)
    findStandards(sheet, results, 2, 2)
    wb.save('ISTD_Results_' + fileName)
